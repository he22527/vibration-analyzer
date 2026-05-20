#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""產生完整計算過程示範 Excel — 所有分頁 + 單位轉換，儲存格皆為公式"""
import pandas as pd, numpy as np, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

# ═══════ 常數 ═══════
ROOT = r"C:\Users\AC717\背景振動量測數據"
FREQ_BANDS = [
    (1,"1 Hz"),(1.25,"1.25 Hz"),(1.6,"1.6 Hz"),(2,"2 Hz"),(2.5,"2.5 Hz"),
    (3.15,"3.15 Hz"),(4,"4 Hz"),(5,"5 Hz"),(6.3,"6.3 Hz"),(8,"8 Hz"),
    (10,"10 Hz"),(12.5,"12.5 Hz"),(16,"16 Hz"),(20,"20 Hz"),(25,"25 Hz"),
    (31.5,"31.5 Hz"),(40,"40 Hz"),(50,"50 Hz"),(63,"63 Hz"),(80,"80 Hz"),
    (100,"100 Hz"),
]
VC_BANDS_IDX = [i for i,(fv,_) in enumerate(FREQ_BANDS) if 4 <= fv <= 80]
AXES = ["X","Y","Z"]
STATS_NAMES = ["Leq","L5","L10","L50","L90","L95","Lmax"]
VC_CURVES = [("工廠 (ISO)",803.2,90),("辦公室 (ISO)",402.6,84),("住宅 (ISO)",201.8,78),
             ("手術室 (ISO)",101.1,72),("VC-A",50.7,66),("VC-B",25.4,60),
             ("VC-C",12.7,54),("VC-D",6.38,48),("VC-E",3.20,42)]
VC_ORDER = list(reversed(VC_CURVES))

# ═══════ 讀取數據 ═══════
print("讀取 .rnd 檔案...")
files = sorted(glob.glob(os.path.join(ROOT,"0416","*.rnd")))
needed = [f"{ax}_{fn}" for _,fn in FREQ_BANDS for ax in AXES]
dfs = []
for f in files:
    try:
        df = pd.read_csv(f, skiprows=1, na_values=["--","UN","OL",""], dtype=str, low_memory=False)
        df.columns = [c.strip() for c in df.columns]
        df["Start Time"] = pd.to_datetime(df["Start Time"].str.strip(),
                                           format="%Y/%m/%d %H:%M:%S.%f", errors="coerce")
        df.dropna(subset=["Start Time"], inplace=True)
        mask = (df["Start Time"] >= "2026-04-16 10:00:00") & (df["Start Time"] <= "2026-04-16 10:00:10.9")
        sub = df.loc[mask]
        if len(sub) > 0: dfs.append(sub)
    except: pass

df = pd.concat(dfs, ignore_index=True)
df.sort_values("Start Time", inplace=True)
df.drop_duplicates(subset=["Start Time"], keep="first", inplace=True)
df.reset_index(drop=True, inplace=True)
for c in df.columns:
    if c != "Start Time" and c in needed:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
keep = ["Start Time"] + [c for c in needed if c in df.columns]
df = df[keep]
N = len(df)
N_BANDS = len(FREQ_BANDS)
N_1S = N // 10
print(f"共 {N} 筆 (0.1s), {N_1S} 秒")

# ═══════ Styles ═══════
hf = Font(name="Microsoft JhengHei", bold=True, size=10, color="FFFFFF")
hfill = PatternFill("solid", fgColor="2C5282")
cf = Font(name="Microsoft JhengHei", bold=True, size=10, color="1A365D")
bf = Font(name="Microsoft JhengHei", size=10)
vf = Font(name="Consolas", size=10)
vfb = Font(name="Consolas", size=10, bold=True, color="C53030")
codef = Font(name="Consolas", size=9, color="6B46C1")
gf = Font(name="Microsoft JhengHei", bold=True, size=10, color="1B5E20")
tf = Font(name="Microsoft JhengHei", bold=True, size=13, color="1A365D")
sf = Font(name="Microsoft JhengHei", bold=True, size=11, color="2C5282")
nf = Font(name="Microsoft JhengHei", size=9, color="718096", italic=True)
thin = Side(style="thin", color="CBD5E0")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
ct = Alignment(horizontal="center", vertical="center")
cw = Alignment(horizontal="center", vertical="center", wrap_text=True)
wp = Alignment(wrap_text=True, vertical="center")
alt = PatternFill("solid", fgColor="F0F4F8")

def hdr_row(ws, r, texts):
    for c, t in enumerate(texts, 1):
        cell = ws.cell(row=r, column=c, value=t)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, cw, bdr

def scell(ws, r, c, val, font=vf, align=ct, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font, cell.alignment, cell.border = font, align, bdr
    if fmt: cell.number_format = fmt
    return cell

def title(ws, r, text, span=12):
    ws.cell(row=r, column=1, value=text).font = tf
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    return r+1

def subtitle(ws, r, text, span=12):
    ws.cell(row=r, column=1, value=text).font = sf
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    return r+1

wb = Workbook()

# ═══════════════════════════════════════════════════════════
# Sheet 1: 原始數據_01s — 110 筆 raw X,Y,Z + XYZ 公式
# ═══════════════════════════════════════════════════════════
print("Sheet 1: 原始數據_01s")
ws1 = wb.active; ws1.title = "原始數據_01s"
SN1 = ws1.title

r = 1
r = title(ws1, r, "原始數據（0.1 秒 / 筆）— X, Y, Z 加速度 + XYZ 三軸合成公式", N_BANDS*4+4)
r = subtitle(ws1, r, "量測期間：2026/04/16 10:00:00.0 ~ 10:00:10.9，共 110 筆，21 頻段", N_BANDS*4+4)
r += 1
HDR1_ROW = r
hdrs = ["筆數","時間"]
for fv, fn in FREQ_BANDS:
    hdrs += [f"X_{fn}", f"Y_{fn}", f"Z_{fn}", f"XYZ_{fn}"]
hdr_row(ws1, r, hdrs)
r += 1
DATA1_START = r

for i in range(N):
    ts = df["Start Time"].iloc[i]
    t_str = ts.strftime("%H:%M:%S.") + f"{ts.microsecond//100000}"
    scell(ws1, r, 1, i+1, cf)
    scell(ws1, r, 2, t_str, vf)
    for bi, (fv, fn) in enumerate(FREQ_BANDS):
        base_col = 3 + bi*4
        for ai, ax in enumerate(AXES):
            col_name = f"{ax}_{fn}"
            val = float(df[col_name].iloc[i]) if col_name in df.columns else 0.0
            scell(ws1, r, base_col+ai, val, vf, fmt="0.00000")
        xc, yc, zc = CL(base_col), CL(base_col+1), CL(base_col+2)
        scell(ws1, r, base_col+3, f"=SQRT({xc}{r}^2+{yc}{r}^2+{zc}{r}^2)", vfb, fmt="0.00000000")
    if i % 2 == 1:
        for c in range(1, len(hdrs)+1):
            ws1.cell(row=r, column=c).fill = alt
    r += 1

DATA1_END = r - 1
ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 14
for c in range(3, len(hdrs)+1):
    ws1.column_dimensions[CL(c)].width = 13
ws1.freeze_panes = f"C{HDR1_ROW+1}"

# XYZ column index for each band (1-based)
XYZ_COL1 = {bi: 3 + bi*4 + 3 for bi in range(N_BANDS)}

# ═══════════════════════════════════════════════════════════
# Sheet 2: RMS聚合_1s — 每 10 筆 → 1 筆 RMS
# ═══════════════════════════════════════════════════════════
print("Sheet 2: RMS聚合_1s")
ws2 = wb.create_sheet("RMS聚合_1s")
SN2 = ws2.title

r = 1
r = title(ws2, r, "RMS 聚合（0.1s → 1s）— 每 10 筆做 √(Σaᵢ²/10)", N_BANDS+4)
r = subtitle(ws2, r, f"引用 '{SN1}' 工作表之 XYZ 合成欄位", N_BANDS+4)
r += 1
HDR2_ROW = r
hdrs2 = ["秒數","起始時間"] + [fn for _,fn in FREQ_BANDS]
hdr_row(ws2, r, hdrs2)
r += 1
DATA2_START = r

for s in range(N_1S):
    row0 = DATA1_START + s*10
    row9 = DATA1_START + s*10 + 9
    ts = df["Start Time"].iloc[s*10]
    t_str = ts.strftime("%H:%M:%S")
    scell(ws2, r, 1, s+1, cf)
    scell(ws2, r, 2, t_str, vf)
    for bi in range(N_BANDS):
        xyz_c = CL(XYZ_COL1[bi])
        scell(ws2, r, 3+bi,
              f"=SQRT(SUMSQ('{SN1}'!{xyz_c}{row0}:{xyz_c}{row9})/10)",
              vfb, fmt="0.00000000")
    if s % 2 == 1:
        for c in range(1, len(hdrs2)+1):
            ws2.cell(row=r, column=c).fill = alt
    r += 1

DATA2_END = r - 1
ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 12
for c in range(3, len(hdrs2)+1):
    ws2.column_dimensions[CL(c)].width = 14
ws2.freeze_panes = f"C{HDR2_ROW+1}"

# ═══════════════════════════════════════════════════════════
# Sheet 3: 統計量_01s
# ═══════════════════════════════════════════════════════════
print("Sheet 3: 統計量_01s")
ws3 = wb.create_sheet("統計量_01s")

r = 1
r = title(ws3, r, "統計量（基於 0.1 秒原始數據 XYZ 合成值）")
r = subtitle(ws3, r, f"引用 '{SN1}' 工作表 XYZ 欄位，共 {N} 筆")
r += 1
HDR3_ROW = r
hdr_row(ws3, r, ["頻段","頻率(Hz)"] + STATS_NAMES)
r += 1

for bi, (fv, fn) in enumerate(FREQ_BANDS):
    xyz_c = CL(XYZ_COL1[bi])
    rng = f"'{SN1}'!{xyz_c}{DATA1_START}:{xyz_c}{DATA1_END}"
    scell(ws3, r, 1, fn, cf)
    scell(ws3, r, 2, fv, vf)
    formulas = [
        f"=SQRT(SUMSQ({rng})/COUNT({rng}))",
        f"=PERCENTILE({rng},0.95)",
        f"=PERCENTILE({rng},0.90)",
        f"=PERCENTILE({rng},0.50)",
        f"=PERCENTILE({rng},0.10)",
        f"=PERCENTILE({rng},0.05)",
        f"=MAX({rng})",
    ]
    for si, fm in enumerate(formulas):
        scell(ws3, r, 3+si, fm, vfb, fmt="0.00000000")
    if bi % 2 == 1:
        for c in range(1, 10):
            ws3.cell(row=r, column=c).fill = alt
    r += 1

STAT01_DATA_START = HDR3_ROW + 1
STAT01_DATA_END = r - 1
ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 10
for c in range(3, 10):
    ws3.column_dimensions[CL(c)].width = 18

# ═══════════════════════════════════════════════════════════
# Sheet 4: 統計量_1s
# ═══════════════════════════════════════════════════════════
print("Sheet 4: 統計量_1s")
ws4 = wb.create_sheet("統計量_1s")

r = 1
r = title(ws4, r, "統計量（基於 1 秒 RMS 聚合值）")
r = subtitle(ws4, r, f"引用 '{SN2}' 工作表，共 {N_1S} 筆")
r += 1
hdr_row(ws4, r, ["頻段","頻率(Hz)"] + STATS_NAMES)
r += 1

for bi, (fv, fn) in enumerate(FREQ_BANDS):
    bc = CL(3 + bi)
    rng = f"'{SN2}'!{bc}{DATA2_START}:{bc}{DATA2_END}"
    scell(ws4, r, 1, fn, cf)
    scell(ws4, r, 2, fv, vf)
    formulas = [
        f"=SQRT(SUMSQ({rng})/COUNT({rng}))",
        f"=PERCENTILE({rng},0.95)",
        f"=PERCENTILE({rng},0.90)",
        f"=PERCENTILE({rng},0.50)",
        f"=PERCENTILE({rng},0.10)",
        f"=PERCENTILE({rng},0.05)",
        f"=MAX({rng})",
    ]
    for si, fm in enumerate(formulas):
        scell(ws4, r, 3+si, fm, vfb, fmt="0.00000000")
    if bi % 2 == 1:
        for c in range(1, 10):
            ws4.cell(row=r, column=c).fill = alt
    r += 1

ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 10
for c in range(3, 10):
    ws4.column_dimensions[CL(c)].width = 18

# ═══════════════════════════════════════════════════════════
# Sheet 5: 時間序列 — 寬頻合成
# ═══════════════════════════════════════════════════════════
print("Sheet 5: 時間序列")
ws5 = wb.create_sheet("時間序列")

r = 1
r = title(ws5, r, "時間序列 — 寬頻合成 BB = √(Σ band_XYZ²)", N_BANDS+5)
r += 1

# 0.1s 寬頻
r = subtitle(ws5, r, "0.1 秒寬頻合成")
hdr_row(ws5, r, ["筆數","時間","BB 公式 = SQRT(Σ XYZ_band²)","BB (m/s²)"])
r += 1
TS01_START = r

for i in range(N):
    ts = df["Start Time"].iloc[i]
    t_str = ts.strftime("%H:%M:%S.") + f"{ts.microsecond//100000}"
    scell(ws5, r, 1, i+1, cf)
    scell(ws5, r, 2, t_str, vf)
    # BB = SQRT(sum of all XYZ^2)
    parts = []
    for bi in range(N_BANDS):
        xyz_c = CL(XYZ_COL1[bi])
        parts.append(f"'{SN1}'!{xyz_c}{DATA1_START+i}^2")
    formula = "=SQRT(" + "+".join(parts) + ")"
    scell(ws5, r, 3, formula[:200]+"…" if len(formula)>200 else "", nf, align=wp)
    scell(ws5, r, 4, formula, vfb, fmt="0.00000000")
    if i % 2 == 1:
        for c in range(1, 5): ws5.cell(row=r, column=c).fill = alt
    r += 1

r += 1
# 1s 寬頻
r = subtitle(ws5, r, "1 秒寬頻合成")
hdr_row(ws5, r, ["秒數","時間","BB (m/s²)"])
r += 1

for s in range(N_1S):
    ts = df["Start Time"].iloc[s*10]
    scell(ws5, r, 1, s+1, cf)
    scell(ws5, r, 2, ts.strftime("%H:%M:%S"), vf)
    parts = []
    for bi in range(N_BANDS):
        bc = CL(3+bi)
        parts.append(f"'{SN2}'!{bc}{DATA2_START+s}^2")
    scell(ws5, r, 3, "=SQRT("+"+".join(parts)+")", vfb, fmt="0.00000000")
    if s % 2 == 1:
        for c in range(1, 4): ws5.cell(row=r, column=c).fill = alt
    r += 1

ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 14
ws5.column_dimensions["C"].width = 60
ws5.column_dimensions["D"].width = 18

# ═══════════════════════════════════════════════════════════
# Sheet 6: 頻率分析 — 各頻段 Leq 比較
# ═══════════════════════════════════════════════════════════
print("Sheet 6: 頻率分析")
ws6 = wb.create_sheet("頻率分析")
SN3 = "統計量_01s"
SN4 = "統計量_1s"

r = 1
r = title(ws6, r, "頻率分析 — 各頻段 Leq 比較（0.1s vs 1s）")
r += 1
hdr_row(ws6, r, ["頻段","頻率(Hz)","Leq_0.1s (m/s²)","Leq_1s (m/s²)","差異(%)"])
r += 1

for bi, (fv, fn) in enumerate(FREQ_BANDS):
    scell(ws6, r, 1, fn, cf)
    scell(ws6, r, 2, fv, vf)
    leq01_ref = f"'{SN3}'!C{STAT01_DATA_START+bi}"
    leq1s_ref = f"'{SN4}'!C{STAT01_DATA_START+bi}"
    scell(ws6, r, 3, f"={leq01_ref}", vfb, fmt="0.00000000")
    scell(ws6, r, 4, f"={leq1s_ref}", vfb, fmt="0.00000000")
    scell(ws6, r, 5, f"=IF(C{r}=0,0,(D{r}-C{r})/C{r}*100)", vf, fmt="0.00")
    if bi % 2 == 1:
        for c in range(1, 6): ws6.cell(row=r, column=c).fill = alt
    r += 1

ws6.column_dimensions["A"].width = 10
ws6.column_dimensions["B"].width = 10
for c in range(3, 6):
    ws6.column_dimensions[CL(c)].width = 20

# ═══════════════════════════════════════════════════════════
# Sheet 7: VC曲線 — 速度轉換 + 等級評定
# ═══════════════════════════════════════════════════════════
print("Sheet 7: VC曲線")
ws7 = wb.create_sheet("VC曲線")

r = 1
r = title(ws7, r, "VC 曲線評定 — 加速度→速度轉換 + 等級判定", 14)
r = subtitle(ws7, r, "評估頻段 4~80 Hz，速度 v = a/(2πf)，單位 μm/s", 14)
r += 1

# 7a: 速度轉換表
r = subtitle(ws7, r, "一、各頻段加速度 Leq → 速度 Leq (μm/s)")
hdr_row(ws7, r, ["頻段","頻率(Hz)","Leq 加速度 (m/s²)","公式: v=a/(2πf)×10⁶","Leq 速度 (μm/s)"])
r += 1
VEL_START = r

for idx in VC_BANDS_IDX:
    fv, fn = FREQ_BANDS[idx]
    leq_ref = f"'{SN3}'!C{STAT01_DATA_START+idx}"
    scell(ws7, r, 1, fn, cf)
    scell(ws7, r, 2, fv, vf)
    scell(ws7, r, 3, f"={leq_ref}", vf, fmt="0.0000E+00")
    scell(ws7, r, 4, f"=C{r}/(2*PI()*B{r})*1000000", codef)
    scell(ws7, r, 5, f"=C{r}/(2*PI()*B{r})*1000000", vfb, fmt="0.0000")
    r += 1

VEL_END = r - 1
r += 1

# 7b: VC 曲線限值表
r = subtitle(ws7, r, "二、VC 曲線限值 (μm/s)", 14)
vc_hdr = ["頻段","頻率(Hz)"] + [name for name,_,_ in VC_CURVES]
hdr_row(ws7, r, vc_hdr)
r += 1
VC_LIM_START = r

for idx in VC_BANDS_IDX:
    fv, fn = FREQ_BANDS[idx]
    scell(ws7, r, 1, fn, cf)
    scell(ws7, r, 2, fv, vf)
    for gi, (gname, flat, db) in enumerate(VC_CURVES):
        if fv >= 8:
            scell(ws7, r, 3+gi, flat, vf, fmt="0.00")
        else:
            steps = round(np.log2(8.0/fv)*3)
            scell(ws7, r, 3+gi, f"={flat}*10^({steps}*2/20)", vf, fmt="0.00")
    r += 1

VC_LIM_END = r - 1
r += 1

# 7c: 等級評定
r = subtitle(ws7, r, "三、VC 等級評定（Leq 速度 vs 曲線限值）", 14)
vc_eval_hdr = ["頻段","Leq速度(μm/s)"] + [name for name,_,_ in VC_CURVES]
hdr_row(ws7, r, vc_eval_hdr)
r += 1

for ki, idx in enumerate(VC_BANDS_IDX):
    vel_row = VEL_START + ki
    lim_row = VC_LIM_START + ki
    scell(ws7, r, 1, FREQ_BANDS[idx][1], cf)
    scell(ws7, r, 2, f"=E{vel_row}", vfb, fmt="0.0000")
    for gi in range(len(VC_CURVES)):
        lim_c = CL(3+gi)
        scell(ws7, r, 3+gi,
              f'=IF(B{r}<={lim_c}{lim_row},"✓","✗")', vf)
    r += 1

r += 1
r = subtitle(ws7, r, "四、等級判定結果")
ws7.cell(row=r, column=1, value="判定邏輯").font = cf
ws7.cell(row=r, column=2, value="從 VC-E 起逐級檢查，所有頻段皆 ≤ 限值 → 該等級通過").font = bf
r += 1

# 手動計算 VC grade for display
for sk_label, pct_col_idx in [("Leq", 0)]:
    band_vels = []
    for idx in VC_BANDS_IDX:
        fv, fn = FREQ_BANDS[idx]
        a_leq = float(np.sqrt(np.mean(
            np.nan_to_num(np.sqrt(
                sum(df[f"{ax}_{fn}"].values**2 for ax in AXES if f"{ax}_{fn}" in df.columns)
            ))**2)))
        v = a_leq / (2*np.pi*fv) * 1e6
        band_vels.append((fv, v))
    grade = "超過工廠 (ISO)"
    for gname, flat, db in reversed(VC_CURVES):
        def _vc(f):
            if f >= 8: return flat
            return flat * 10**(round(np.log2(8.0/f)*3)*2/20)
        if all(vel <= _vc(fv) for fv, vel in band_vels):
            grade = gname; break
    peak = max(v for _, v in band_vels)
    scell(ws7, r, 1, "VC 等級 (Leq)", gf)
    scell(ws7, r, 2, grade, gf)
    scell(ws7, r, 3, f"峰值速度: {peak:.4f} μm/s", vf)
    r += 1

for c in range(1, 14):
    ws7.column_dimensions[CL(c)].width = 14
ws7.column_dimensions["A"].width = 10
ws7.column_dimensions["D"].width = 28

# ═══════════════════════════════════════════════════════════
# Sheet 8: 環境振動
# ═══════════════════════════════════════════════════════════
print("Sheet 8: 環境振動")
ws8 = wb.create_sheet("環境振動")

r = 1
r = title(ws8, r, "環境振動評估表")
r = subtitle(ws8, r, "期間: 2026/04/16 10:00:00~10:00:10 (僅 11 秒示範)")
r += 1

ws8.cell(row=r, column=1, value="說明").font = cf
ws8.cell(row=r, column=2, value=(
    "環境振動分頁將量測數據依小時分段，計算各時段的速度位準統計量 (Lveq, Lvmax, Lv5...Lv95)。"
    "本範例僅 11 秒數據，無法分段，故直接對全部數據計算。"
    "速度位準 dB = 20·log₁₀(v_μm/s / 0.0254)，ref 2.54×10⁻⁸ m/s。"
)).font = nf
ws8.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
r += 2

hdr_row(ws8, r, ["時段","Lveq (dB)","Lvmax (dB)","Lv5 (dB)","Lv10 (dB)",
                  "Lv50 (dB)","Lv90 (dB)","Lv95 (dB)"])
r += 1

# 寬頻速度 dB: 先算寬頻加速度，再轉速度，再轉 dB
# 用代表頻段的寬頻 Leq 計算 (簡化：用 VC曲線 sheet 的速度值)
ws8.cell(row=r, column=1, value="10:00").font = cf
ws8.cell(row=r, column=1).border = bdr
ws8.cell(row=r, column=1).alignment = ct
note_text = "（需完整 1 小時數據，此處僅為公式示範框架）"
for ci in range(2, 9):
    ws8.cell(row=r, column=ci, value="—").font = vf
    ws8.cell(row=r, column=ci).border = bdr
    ws8.cell(row=r, column=ci).alignment = ct
r += 2

ws8.cell(row=r, column=1, value="公式說明").font = sf
r += 1
env_desc = [
    ("Lveq", "20·log₁₀(v_Leq / 0.0254)，v_Leq 為寬頻速度 Leq (μm/s)"),
    ("Lvmax", "20·log₁₀(v_Lmax / 0.0254)"),
    ("Lv5~Lv95", "20·log₁₀(v_Ln / 0.0254)，Ln 為速度的超越百分位"),
    ("ref 值", "0.0254 μm/s = 2.54×10⁻⁸ m/s = 10⁻⁶ in/s"),
]
for label, desc in env_desc:
    scell(ws8, r, 1, label, cf)
    scell(ws8, r, 2, desc, bf, align=wp)
    ws8.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1

ws8.column_dimensions["A"].width = 12
for c in range(2, 9):
    ws8.column_dimensions[CL(c)].width = 14

# ═══════════════════════════════════════════════════════════
# Sheet 9~13: 單位轉換
# ═══════════════════════════════════════════════════════════
# Helper: 用 統計量_01s 的 Leq 值做轉換示範
SN_S01 = "統計量_01s"

def make_conv_sheet(ws, title_text, sub_text, conv_hdr, conv_formula_fn, fmt_out="0.000000"):
    """建立單位轉換工作表"""
    r = 1
    r = title(ws, r, title_text)
    r = subtitle(ws, r, sub_text)
    r += 1
    hdr_row(ws, r, conv_hdr)
    r += 1
    for bi, (fv, fn) in enumerate(FREQ_BANDS):
        scell(ws, r, 1, fn, cf)
        scell(ws, r, 2, fv, vf)
        leq_ref = f"'{SN_S01}'!C{STAT01_DATA_START+bi}"
        scell(ws, r, 3, f"={leq_ref}", vf, fmt="0.0000E+00")
        formulas = conv_formula_fn(r, fv)
        for fi, (fm, ft) in enumerate(formulas):
            scell(ws, r, 4+fi, fm, vfb if fi == len(formulas)-1 else codef, fmt=ft)
        if bi % 2 == 1:
            for c in range(1, 4+len(formulas)+1):
                ws.cell(row=r, column=c).fill = alt
        r += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    for c in range(4, 10):
        ws.column_dimensions[CL(c)].width = 28

# 9. 轉換_g
print("Sheet 9: 轉換_g")
ws9 = wb.create_sheet("轉換_g")
make_conv_sheet(ws9,
    "加速度單位轉換 — m/s² → g",
    "g = a / 9.80665，1 g = 9.80665 m/s²",
    ["頻段","頻率(Hz)","Leq (m/s²)","公式: a/9.80665","結果 (g)"],
    lambda r, fv: [
        (f"=C{r}/9.80665", "0.0000E+00"),
    ])

# 10. 轉換_dB加速度
print("Sheet 10: 轉換_dB加速度")
ws10 = wb.create_sheet("轉換_dB加速度")
make_conv_sheet(ws10,
    "加速度位準 dB — ref 10⁻⁶ m/s²",
    "dB = 20·log₁₀(a / 10⁻⁶)，基準 a₀ = 10⁻⁶ m/s²（ISO 1683）",
    ["頻段","頻率(Hz)","Leq (m/s²)","公式: 20*LOG10(a/1E-6)","結果 (dB)"],
    lambda r, fv: [
        (f"=20*LOG10(C{r}/0.000001)", "0.00"),
    ])

# 11. 轉換_速度ms
print("Sheet 11: 轉換_速度ms")
ws11 = wb.create_sheet("轉換_速度ms")
make_conv_sheet(ws11,
    "加速度 → 速度 m/s",
    "v = a / (2π·f)，頻率 f 為各頻段中心頻率",
    ["頻段","頻率(Hz)","Leq (m/s²)","公式: a/(2*PI()*f)","v (m/s)","v (μm/s)"],
    lambda r, fv: [
        (f"=C{r}/(2*PI()*B{r})", "0.0000E+00"),
        (f"=D{r}*1000000", "0.0000"),
    ])

# 12. 轉換_ips
print("Sheet 12: 轉換_ips")
ws12 = wb.create_sheet("轉換_ips")
make_conv_sheet(ws12,
    "速度 → in/s（英制）",
    "ips = v_m/s × 39.3701，μin/s = μm/s / 0.0254",
    ["頻段","頻率(Hz)","Leq (m/s²)","v (m/s) = a/(2πf)","v (in/s)","v (μin/s)"],
    lambda r, fv: [
        (f"=C{r}/(2*PI()*B{r})", "0.0000E+00"),
        (f"=D{r}*39.3701", "0.0000E+00"),
        (f"=D{r}*39.3701*1000000", "0.00"),
    ])

# 13. 轉換_dB速度
print("Sheet 13: 轉換_dB速度")
ws13 = wb.create_sheet("轉換_dB速度")
make_conv_sheet(ws13,
    "速度位準 dB — ref 2.54×10⁻⁸ m/s",
    "dB = 20·log₁₀(v / 2.54×10⁻⁸)，等效於 20·log₁₀(v_μm/s / 0.0254)",
    ["頻段","頻率(Hz)","Leq (m/s²)","v (m/s) = a/(2πf)","v (μm/s)","dB = 20·LOG10(v_μm/s/0.0254)"],
    lambda r, fv: [
        (f"=C{r}/(2*PI()*B{r})", "0.0000E+00"),
        (f"=D{r}*1000000", "0.0000"),
        (f"=20*LOG10(E{r}/0.0254)", "0.00"),
    ])

# ═══════════════════════════════════════════════════════════
# 儲存
# ═══════════════════════════════════════════════════════════
out = r"S:\noise\Vibration&Noise\project\15002-(寶山用地)第2期擴建-再生水輸水管線工程設計及監造工作(設計部分)\2.工作區\22527\計算過程示範.xlsx"
wb.save(out)
print(f"\nOK: {out}")
print(f"共 {len(wb.sheetnames)} 個工作表: {', '.join(wb.sheetnames)}")
