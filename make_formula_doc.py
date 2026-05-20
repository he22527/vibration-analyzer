#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""產生公式說明 Excel 檔"""
import json, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ═══════ Styles ═══════
hdr_font = Font(name="Microsoft JhengHei", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2C5282")
cat_font = Font(name="Microsoft JhengHei", bold=True, size=10, color="1A365D")
cat_fill = PatternFill("solid", fgColor="E2E8F0")
body_font = Font(name="Microsoft JhengHei", size=10)
code_font = Font(name="Consolas", size=9, color="6B46C1")
note_font = Font(name="Microsoft JhengHei", size=9, color="718096", italic=True)
title_font = Font(name="Microsoft JhengHei", bold=True, size=13, color="1A365D")
sub_font = Font(name="Microsoft JhengHei", bold=True, size=11, color="2C5282")
val_font = Font(name="Consolas", size=10)
bold_val = Font(name="Consolas", size=10, bold=True, color="C53030")
green_font = Font(name="Microsoft JhengHei", bold=True, size=10, color="1B5E20")
db_font = Font(name="Consolas", size=10, bold=True, color="2C5282")
thin = Side(style="thin", color="CBD5E0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(horizontal="center", vertical="center")
cwrap = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = cwrap
        cell.border = border


def style_row(ws, row, cols, fonts=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fonts[c - 1] if fonts else body_font
        cell.alignment = wrap
        cell.border = border


def title_cell(ws, row, text, span=8):
    ws.cell(row=row, column=1, value=text).font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def subtitle_cell(ws, row, text, span=8):
    ws.cell(row=row, column=1, value=text).font = sub_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


# ═══════════════════════════════════════
# Sheet 1: 公式總覽
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "公式總覽"

headers = ["編號", "分類", "公式名稱", "數學公式", "Python 函式 / 程式碼", "說明"]
ws1.append(headers)
style_hdr(ws1, 1, 6)

formulas = [
    ("統計量", "Leq（能量均方根 RMS）",
     "Leq = √(1/N × Σaᵢ²)", "np.sqrt(np.mean(v**2))",
     "量測期間所有數據的能量等效值（均方根）"),
    ("統計量", "L5",
     "L5 = P₉₅(v)", "np.percentile(v, 95)",
     "量測期間有 5% 的數據 ≥ 此值\n→ 取排序後第 95 百分位數"),
    ("統計量", "L10",
     "L10 = P₉₀(v)", "np.percentile(v, 90)",
     "量測期間有 10% 的數據 ≥ 此值\n→ 取排序後第 90 百分位數"),
    ("統計量", "L50（中位數）",
     "L50 = P₅₀(v)", "np.percentile(v, 50)",
     "量測期間有 50% 的數據 ≥ 此值\n→ 中位數"),
    ("統計量", "L90",
     "L90 = P₁₀(v)", "np.percentile(v, 10)",
     "量測期間有 90% 的數據 ≥ 此值\n→ 取排序後第 10 百分位數"),
    ("統計量", "L95",
     "L95 = P₅(v)", "np.percentile(v, 5)",
     "量測期間有 95% 的數據 ≥ 此值\n→ 取排序後第 5 百分位數"),
    ("統計量", "Lmax（最大值）",
     "Lmax = max(v)", "np.max(v)",
     "量測期間的最大值（0% 超越）"),
    ("統計量", "排序關係",
     "L5 > L10 > L50 > L90 > L95", "—",
     "超越機率越大 → 百分位越低 → 數值越小"),
    ("單位轉換", "加速度 → 速度",
     "v = a / (2π·f)", "df[col] / (2.0 * np.pi * f_val)",
     "將加速度 a (m/s²) 轉換為速度 v (m/s)\nf = 頻率 (Hz)"),
    ("單位轉換", "加速度位準 dB",
     "dB = 20·log₁₀(a / a₀)\na₀ = 10⁻⁶ m/s²", "20 * np.log10(val / 1e-6)",
     "加速度級，基準 a₀ = 10⁻⁶ m/s²\n（ISO 1683）"),
    ("單位轉換", "速度位準 dB",
     "dB = 20·log₁₀(v / v₀)\nv₀ = 2.54×10⁻⁸ m/s", "20 * np.log10(val / 2.54e-8)",
     "速度級，基準 v₀ = 2.54×10⁻⁸ m/s\n（= 10⁻⁶ in/s）"),
    ("單位轉換", "m/s² → g",
     "g = a / 9.80665", "val * (1 / 9.80665)",
     "重力加速度；1 g = 9.80665 m/s²"),
    ("單位轉換", "m/s² → Gal",
     "Gal = a × 100", "val * 100",
     "1 Gal = 0.01 m/s² = 1 cm/s²"),
    ("單位轉換", "m/s → mm/s",
     "mm/s = v × 1000", "val * 1000",
     "公制速度"),
    ("單位轉換", "m/s → ips (in/s)",
     "ips = v × 39.3701", "val * 39.3701",
     "英制速度"),
    ("單位轉換", "μm/s → μin/s",
     "μin/s = μm/s / 0.0254", "val_ums / 0.0254",
     "微米每秒轉微英寸每秒\n1 μin = 0.0254 μm"),
    ("XYZ 合成", "三軸合成（逐筆逐頻段）",
     "XYZ = √(X² + Y² + Z²)", "np.sqrt(np.sum(mat**2, axis=1))",
     "逐時間步、逐頻段合成三軸向量和"),
    ("XYZ 合成", "寬頻合成",
     "BB = √(Σ bandᵢ²)", "sq += a**2; bb = np.sqrt(sq)",
     "將各頻段值合成為單一寬頻值"),
    ("RMS 聚合", "區間 RMS",
     "RMS = √(1/N × Σaᵢ²)", "np.sqrt(np.mean(a**2))",
     "依分析間距（1秒~60分鐘）\n對原始 0.1s 數據做 RMS 聚合"),
    ("VC 曲線", "平坦段（≥ 8 Hz）",
     "v_limit = flat_limit", "flat_limit（μm/s）",
     "8~80 Hz 等速度段，限值固定不變"),
    ("VC 曲線", "斜率段（4~8 Hz）",
     "v = flat × 10^(steps×2/20)", "flat * 10**(steps*2/20)",
     "4~8 Hz 等加速度段\n每 1/3 八度音帶 +2 dB"),
    ("VC 曲線", "1/3 八度音帶步數",
     "steps = round(log₂(8/f) × 3)", "round(np.log2(8.0/f) * 3)",
     "計算頻率 f 到 8 Hz 的步數\n4Hz→3步、5Hz→2步、6.3Hz→1步"),
    ("VC 曲線", "VC 等級評定",
     "所有頻段 vel ≤ 曲線值\n→ 該等級通過",
     "all(vel <= _vc_curve_val(flat, fv))",
     "從 VC-E 起往上逐級檢查\n首個全部通過的即為等級"),
    ("VC 曲線", "Y軸加速度位準 dB",
     "dB = 20·log₁₀(v_μm/s × 2π·f)",
     "20*np.log10(val_ums*2*np.pi*freq)",
     "速度→加速度位準 dB\nref 10⁻⁶ m/s²"),
    ("VC 曲線", "Y軸速度位準 dB",
     "dB = 20·log₁₀(v_μm/s / 0.0254)",
     "20*np.log10(val_ums / 0.0254)",
     "速度位準 dB\nref 2.54×10⁻⁸ m/s = 10⁻⁶ in/s"),
]

for i, (cat, name, formula, code, desc) in enumerate(formulas, 1):
    ws1.append([i, cat, name, formula, code, desc])
    r = i + 1
    fonts = [body_font, cat_font, body_font, val_font, code_font, body_font]
    style_row(ws1, r, 6, fonts)
    ws1.cell(row=r, column=1).alignment = center
    ws1.cell(row=r, column=2).alignment = center

ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 14
ws1.column_dimensions["C"].width = 26
ws1.column_dimensions["D"].width = 36
ws1.column_dimensions["E"].width = 42
ws1.column_dimensions["F"].width = 38
ws1.sheet_properties.pageSetUpPr = None
ws1.freeze_panes = "A2"

# ═══════════════════════════════════════
# Sheet 2: 範例計算
# ═══════════════════════════════════════
ws2 = wb.create_sheet("範例計算")
r = 1

# ── 一、統計量 ──
r = title_cell(ws2, r, "一、統計量計算範例", 10)
r = subtitle_cell(ws2, r, "原始數據（10 筆加速度值，單位 m/s²）", 10)

data = [3.5, 4.2, 5.1, 3.8, 6.0, 4.5, 3.2, 7.1, 4.8, 5.5]
sorted_data = sorted(data)
hdr = [f"第{i}筆" for i in range(1, 11)]
for c, h in enumerate(hdr, 1):
    ws2.cell(row=r, column=c, value=h)
style_hdr(ws2, r, 10)
r += 1
for c, v in enumerate(data, 1):
    cell = ws2.cell(row=r, column=c, value=v)
    cell.font = val_font
    cell.alignment = center
    cell.border = border
r += 1

r = subtitle_cell(ws2, r, "排序後（升序）", 10)
for c, v in enumerate(sorted_data, 1):
    cell = ws2.cell(row=r, column=c, value=v)
    cell.font = val_font
    cell.alignment = center
    cell.border = border
r += 2

r = subtitle_cell(ws2, r, "統計量計算結果")
stat_hdr = ["統計量", "計算方法", "計算過程", "結果"]
for c, h in enumerate(stat_hdr, 1):
    ws2.cell(row=r, column=c, value=h)
style_hdr(ws2, r, 4)
r += 1

v = np.array(data)
leq = float(np.sqrt(np.mean(v ** 2)))
stats_ex = [
    ("Leq", "RMS = √(Σaᵢ²/N)",
     "√((3.5²+4.2²+…+5.5²)/10) = √(24.153)", f"{leq:.3f}"),
    ("L5", "第 95 百分位數",
     "排序後取 95% 位置（頂端 5%）", f"{float(np.percentile(v, 95)):.3f}"),
    ("L10", "第 90 百分位數",
     "排序後取 90% 位置（頂端 10%）", f"{float(np.percentile(v, 90)):.3f}"),
    ("L50", "第 50 百分位數",
     "排序後取 50% 位置（中位數）", f"{float(np.percentile(v, 50)):.3f}"),
    ("L90", "第 10 百分位數",
     "排序後取 10% 位置（底端 10%）", f"{float(np.percentile(v, 10)):.3f}"),
    ("L95", "第 5 百分位數",
     "排序後取 5% 位置（底端 5%）", f"{float(np.percentile(v, 5)):.3f}"),
    ("Lmax", "最大值", "max(全部數據) = 7.1", "7.100"),
]
for name, method, process, result in stats_ex:
    for c, val in enumerate([name, method, process, result], 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, [cat_font, body_font, val_font, val_font])
    ws2.cell(row=r, column=1).alignment = center
    r += 1

r += 1
ws2.cell(row=r, column=1,
         value=f"排序驗證：{float(np.percentile(v,95)):.3f} > "
               f"{float(np.percentile(v,90)):.3f} > "
               f"{float(np.percentile(v,50)):.3f} > "
               f"{float(np.percentile(v,10)):.3f} > "
               f"{float(np.percentile(v,5)):.3f}  ✓ L5 > L10 > L50 > L90 > L95 成立").font = val_font
r += 2

# ── 二、單位轉換 ──
r = title_cell(ws2, r, "二、單位轉換範例")
r = subtitle_cell(ws2, r, "已知：加速度 a = 0.005 m/s²，頻率 f = 10 Hz")
r += 1

conv_hdr = ["轉換項目", "公式", "代入數值", "結果"]
for c, h in enumerate(conv_hdr, 1):
    ws2.cell(row=r, column=c, value=h)
style_hdr(ws2, r, 4)
r += 1

a_val = 0.005
f_val = 10
vel = a_val / (2 * np.pi * f_val)
convs = [
    ("加速度→速度", "v = a/(2π·f)",
     f"0.005 / (2π×10) = {vel:.4e} m/s", f"{vel * 1e6:.2f} μm/s"),
    ("加速度 dB", "20·log₁₀(a/10⁻⁶)",
     f"20·log₁₀(0.005/10⁻⁶) = 20·log₁₀(5000)", f"{20 * np.log10(a_val / 1e-6):.2f} dB"),
    ("速度 dB", "20·log₁₀(v/2.54×10⁻⁸)",
     f"20·log₁₀({vel:.4e} / 2.54×10⁻⁸)", f"{20 * np.log10(vel / 2.54e-8):.2f} dB"),
    ("m/s² → g", "a / 9.80665",
     "0.005 / 9.80665", f"{a_val / 9.80665:.4e} g"),
    ("m/s² → Gal", "a × 100",
     "0.005 × 100", f"{a_val * 100:.3f} Gal"),
    ("m/s → mm/s", "v × 1000",
     f"{vel:.4e} × 1000", f"{vel * 1000:.5f} mm/s"),
    ("m/s → ips", "v × 39.3701",
     f"{vel:.4e} × 39.3701", f"{vel * 39.3701:.4e} in/s"),
    ("μm/s → μin/s", "μm/s / 0.0254",
     f"{vel * 1e6:.2f} / 0.0254", f"{vel * 1e6 / 0.0254:.1f} μin/s"),
]
for item, formula, calc, result in convs:
    for c, val in enumerate([item, formula, calc, result], 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, [cat_font, body_font, val_font, val_font])
    r += 1
r += 1

# ── 三、XYZ 合成 ──
r = title_cell(ws2, r, "三、XYZ 三軸合成範例")
r = subtitle_cell(ws2, r, "已知：X = 2.0 m/s²、Y = 3.0 m/s²、Z = 4.0 m/s²（同一頻段、同一時步）")
r += 1

xyz = np.sqrt(2 ** 2 + 3 ** 2 + 4 ** 2)
for label, val, fnt in [
    ("公式", "XYZ = √(X² + Y² + Z²)", body_font),
    ("代入", "XYZ = √(2² + 3² + 4²) = √(4 + 9 + 16) = √29", val_font),
    ("結果", f"{xyz:.3f} m/s²", bold_val),
]:
    ws2.cell(row=r, column=1, value=label).font = cat_font
    ws2.cell(row=r, column=1).border = border
    ws2.cell(row=r, column=2, value=val).font = fnt
    ws2.cell(row=r, column=2).border = border
    r += 1
r += 1

# ── 四、VC 曲線 ──
r = title_cell(ws2, r, "四、VC 曲線計算範例")
r = subtitle_cell(ws2, r, "VC-C 曲線（平坦段限值 = 12.7 μm/s，dB = 54）")
r += 1

vc_hdr = ["頻率 (Hz)", "步數 (steps)", "乘數 10^(steps×2/20)",
          "曲線值 (μm/s)", "曲線值 (dB)"]
for c, h in enumerate(vc_hdr, 1):
    ws2.cell(row=r, column=c, value=h)
style_hdr(ws2, r, 5)
r += 1

flat = 12.7
for freq in [4, 5, 6.3, 8, 10, 16, 25, 40, 63, 80]:
    if freq >= 8:
        steps = 0
        mult = 1.0
    else:
        steps = round(np.log2(8.0 / freq) * 3)
        mult = 10 ** (steps * 2 / 20)
    cv = flat * mult
    db = 20 * np.log10(cv / 0.0254)
    for c, val in enumerate([freq, steps, f"{mult:.4f}", f"{cv:.2f}", f"{db:.1f}"], 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = val_font
        cell.alignment = center
        cell.border = border
    r += 1
r += 1

r = subtitle_cell(ws2, r, "VC 等級評定範例：量測峰值速度 = 20 μm/s（@ 10 Hz，平坦段）")
r += 1

grade_hdr = ["檢查等級", "平坦段限值 (μm/s)", "10Hz 曲線值 (μm/s)",
             "20 ≤ 曲線值？", "結果"]
for c, h in enumerate(grade_hdr, 1):
    ws2.cell(row=r, column=c, value=h)
style_hdr(ws2, r, 5)
r += 1

checks = [
    ("VC-E", 3.20, "否 (20 > 3.20)", "✗ 不通過"),
    ("VC-D", 6.38, "否 (20 > 6.38)", "✗ 不通過"),
    ("VC-C", 12.7, "否 (20 > 12.7)", "✗ 不通過"),
    ("VC-B", 25.4, "是 (20 ≤ 25.4)", "✓ 通過 → 等級 = VC-B"),
]
for name, flat_v, compare, result in checks:
    for c, val in enumerate([name, flat_v, flat_v, compare, result], 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = center
        cell.border = border
        cell.font = body_font
    ws2.cell(row=r, column=1).font = cat_font
    ws2.cell(row=r, column=2).font = val_font
    ws2.cell(row=r, column=3).font = val_font
    if "通過 →" in result:
        ws2.cell(row=r, column=5).font = green_font
    r += 1

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 52
ws2.column_dimensions["D"].width = 24
ws2.column_dimensions["E"].width = 28

# ═══════════════════════════════════════
# Sheet 3: VC曲線標準值
# ═══════════════════════════════════════
ws3 = wb.create_sheet("VC曲線標準值")

r3 = 1
ws3.cell(row=r3, column=1,
         value="VC 曲線標準值（μm/s）— Ungar & Gordon (1991)").font = title_font
ws3.merge_cells("A1:P1")
r3 = 2
ws3.cell(row=r3, column=1,
         value="dB 基準值：1×10⁻⁶ in/s = 2.54×10⁻⁸ m/s。"
               "4~8 Hz 每 1/3 八度音帶 +2 dB（等加速度斜率段）。").font = note_font
ws3.merge_cells("A2:P2")
r3 = 3

freqs = [4, 5, 6.3, 8, 10, 12.5, 16, 20, 25, 31.5, 40, 50, 63, 80]
hdr3 = ["場所", "平坦段 dB"] + [f"{f} Hz" for f in freqs]
for c, h in enumerate(hdr3, 1):
    ws3.cell(row=r3, column=c, value=h)
style_hdr(ws3, r3, len(hdr3))
r3 += 1

curves = [
    ("工廠 (ISO)", 803.2, 90),
    ("辦公室 (ISO)", 402.6, 84),
    ("住宅 (ISO)", 201.8, 78),
    ("手術室 (ISO)", 101.1, 72),
    ("VC-A", 50.7, 66),
    ("VC-B", 25.4, 60),
    ("VC-C", 12.7, 54),
    ("VC-D", 6.38, 48),
    ("VC-E", 3.20, 42),
]
for name, flat_v, db_v in curves:
    vals = []
    for fq in freqs:
        if fq >= 8:
            vals.append(flat_v)
        else:
            steps = round(np.log2(8.0 / fq) * 3)
            vals.append(flat_v * 10 ** (steps * 2 / 20))
    row_data = [name, db_v] + [
        round(v, 1) if v >= 10 else round(v, 2) for v in vals
    ]
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r3, column=c, value=val)
        cell.border = border
        cell.alignment = center
        if c == 1:
            cell.font = cat_font
        elif c == 2:
            cell.font = db_font
        else:
            cell.font = val_font
    r3 += 1

r3 += 1
ws3.cell(row=r3, column=1,
         value="VC 曲線標準值（dB，ref 1×10⁻⁶ in/s = 2.54×10⁻⁸ m/s）").font = sub_font
ws3.merge_cells(f"A{r3}:P{r3}")
r3 += 1

for c, h in enumerate(hdr3, 1):
    ws3.cell(row=r3, column=c, value=h)
style_hdr(ws3, r3, len(hdr3))
r3 += 1

for name, flat_v, db_v in curves:
    vals = []
    for fq in freqs:
        if fq >= 8:
            vals.append(db_v)
        else:
            steps = round(np.log2(8.0 / fq) * 3)
            vals.append(db_v + steps * 2)
    row_data = [name, db_v] + vals
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r3, column=c, value=val)
        cell.border = border
        cell.alignment = center
        cell.font = cat_font if c == 1 else val_font
    r3 += 1

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 12
for c in range(3, 17):
    ws3.column_dimensions[get_column_letter(c)].width = 10

# ═══════════════════════════════════════
# Sheet 4: 實測範例
# ═══════════════════════════════════════
with open(r"C:\Users\AC717\背景振動量測數據\sample_result.json", "r", encoding="utf-8") as _f:
    RD = json.load(_f)

ws4 = wb.create_sheet("實測範例")
r4 = 1

# ── 數據概覽 ──
r4 = title_cell(ws4, r4, "實測數據分析範例 — 2026/04/16 10:00~11:00", 10)
r4 = subtitle_cell(ws4, r4, "一、數據概覽", 10)

overview = [
    ("量測時間", f"{RD['time_start']}  ~  {RD['time_end']}"),
    ("量測間距", "0.1 秒 / 筆"),
    ("總筆數", f"{RD['n_rows']:,} 筆  （= 3600 秒 × 10 筆/秒）"),
    ("頻段", "1/3 八度音帶  1 Hz ~ 100 Hz  共 21 頻段"),
    ("軸向", "XYZ 三軸合成  √(X² + Y² + Z²)"),
]
for label, val in overview:
    ws4.cell(row=r4, column=1, value=label).font = cat_font
    ws4.cell(row=r4, column=1).border = border
    ws4.cell(row=r4, column=2, value=val).font = val_font
    ws4.cell(row=r4, column=2).border = border
    ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=6)
    r4 += 1
r4 += 1

# ── 原始數據示例 ──
r4 = subtitle_cell(ws4, r4, "二、原始數據示例（前 20 筆 XYZ 合成加速度）", 10)
for fn in ["10 Hz", "31.5 Hz"]:
    ws4.cell(row=r4, column=1, value=f"頻段 {fn}（m/s²）").font = cat_font
    ws4.cell(row=r4, column=1).border = border
    samples = RD["samples"].get(fn, [])
    for j, sv in enumerate(samples):
        cell = ws4.cell(row=r4, column=2 + j, value=f"{sv:.2e}")
        cell.font = val_font
        cell.alignment = center
        cell.border = border
    r4 += 1
r4 += 1

# ── 加速度統計 (VC評估頻段 4~80 Hz) ──
vc_bands = ["4 Hz","5 Hz","6.3 Hz","8 Hz","10 Hz","12.5 Hz","16 Hz",
            "20 Hz","25 Hz","31.5 Hz","40 Hz","50 Hz","63 Hz","80 Hz"]
STATS = ["Leq","L5","L10","L50","L90","L95","Lmax"]

r4 = subtitle_cell(ws4, r4, "三、各頻段加速度統計（XYZ 合成，單位 m/s²）", 10)
acc_hdr = ["頻段"] + STATS
for c, h in enumerate(acc_hdr, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, len(acc_hdr))
r4 += 1

for fn in vc_bands:
    st = RD["acc_stats"].get(fn, {})
    ws4.cell(row=r4, column=1, value=fn).font = cat_font
    ws4.cell(row=r4, column=1).alignment = center
    ws4.cell(row=r4, column=1).border = border
    for j, s in enumerate(STATS):
        v = st.get(s, None)
        cell = ws4.cell(row=r4, column=2 + j, value=f"{v:.2e}" if v else "—")
        cell.font = val_font
        cell.alignment = center
        cell.border = border
    r4 += 1
r4 += 1

# ── 速度統計 ──
r4 = subtitle_cell(ws4, r4, "四、各頻段速度統計（v = a/(2πf)，單位 μm/s）", 10)
for c, h in enumerate(acc_hdr, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, len(acc_hdr))
r4 += 1

for fn in vc_bands:
    st = RD["vel_stats"].get(fn, {})
    ws4.cell(row=r4, column=1, value=fn).font = cat_font
    ws4.cell(row=r4, column=1).alignment = center
    ws4.cell(row=r4, column=1).border = border
    for j, s in enumerate(STATS):
        v = st.get(s, None)
        cell = ws4.cell(row=r4, column=2 + j, value=f"{v:.4f}" if v else "—")
        cell.font = val_font
        cell.alignment = center
        cell.border = border
    r4 += 1
r4 += 1

# ── 單位轉換示範（用 10 Hz Leq 為例）──
r4 = subtitle_cell(ws4, r4, "五、單位轉換示範（10 Hz 頻段 Leq 為例）", 10)
a_10 = RD["acc_stats"]["10 Hz"]["Leq"]
v_10 = RD["vel_stats"]["10 Hz"]["Leq"]
db_a10 = RD["db_acc_stats"]["10 Hz"]["Leq"]
db_v10 = RD["db_vel_stats"]["10 Hz"]["Leq"]

conv_rows = [
    ("原始 XYZ 加速度 Leq", "—", f"{a_10:.4e} m/s²"),
    ("加速度 → 速度", "v = a/(2π×10)", f"{a_10:.4e} / 62.832 = {a_10/62.832:.4e} m/s = {v_10:.4f} μm/s"),
    ("加速度 dB", "20·log₁₀(a / 10⁻⁶)", f"20·log₁₀({a_10:.4e} / 10⁻⁶) = {db_a10:.2f} dB"),
    ("速度 dB", "20·log₁₀(v_μm/s / 0.0254)", f"20·log₁₀({v_10:.4f} / 0.0254) = {db_v10:.2f} dB"),
    ("μm/s → μin/s", "μm/s / 0.0254", f"{v_10:.4f} / 0.0254 = {v_10/0.0254:.2f} μin/s"),
]
conv_h = ["項目", "公式", "計算過程與結果"]
for c, h in enumerate(conv_h, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, 3)
r4 += 1
for item, formula, result in conv_rows:
    ws4.cell(row=r4, column=1, value=item).font = cat_font
    ws4.cell(row=r4, column=1).border = border
    ws4.cell(row=r4, column=2, value=formula).font = body_font
    ws4.cell(row=r4, column=2).border = border
    ws4.cell(row=r4, column=3, value=result).font = val_font
    ws4.cell(row=r4, column=3).border = border
    r4 += 1
r4 += 1

# ── dB 統計表 ──
r4 = subtitle_cell(ws4, r4, "六、各頻段 dB 統計（加速度 dB，ref 10⁻⁶ m/s²）", 10)
for c, h in enumerate(acc_hdr, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, len(acc_hdr))
r4 += 1

for fn in vc_bands:
    st = RD["db_acc_stats"].get(fn, {})
    ws4.cell(row=r4, column=1, value=fn).font = cat_font
    ws4.cell(row=r4, column=1).alignment = center
    ws4.cell(row=r4, column=1).border = border
    for j, s in enumerate(STATS):
        v = st.get(s, None)
        cell = ws4.cell(row=r4, column=2 + j, value=f"{v:.2f}" if v is not None else "—")
        cell.font = val_font
        cell.alignment = center
        cell.border = border
    r4 += 1
r4 += 1

r4 = subtitle_cell(ws4, r4, "七、各頻段 dB 統計（速度 dB，ref 2.54×10⁻⁸ m/s）", 10)
for c, h in enumerate(acc_hdr, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, len(acc_hdr))
r4 += 1

for fn in vc_bands:
    st = RD["db_vel_stats"].get(fn, {})
    ws4.cell(row=r4, column=1, value=fn).font = cat_font
    ws4.cell(row=r4, column=1).alignment = center
    ws4.cell(row=r4, column=1).border = border
    for j, s in enumerate(STATS):
        v = st.get(s, None)
        cell = ws4.cell(row=r4, column=2 + j, value=f"{v:.2f}" if v is not None else "—")
        cell.font = val_font
        cell.alignment = center
        cell.border = border
    r4 += 1
r4 += 1

# ── VC 等級評定 ──
r4 = subtitle_cell(ws4, r4, "八、VC 等級評定（4~80 Hz 速度 μm/s）", 10)
r4 = subtitle_cell(ws4, r4, "以 Leq 為例，逐頻段速度值 vs 各等級曲線限值", 10)

VC_CURVES_L = {"工廠 (ISO)": 803.2, "辦公室 (ISO)": 402.6, "住宅 (ISO)": 201.8,
               "手術室 (ISO)": 101.1, "VC-A": 50.7, "VC-B": 25.4,
               "VC-C": 12.7, "VC-D": 6.38, "VC-E": 3.20}
VC_ORDER_L = ["VC-E","VC-D","VC-C","VC-B","VC-A","手術室 (ISO)","住宅 (ISO)","辦公室 (ISO)","工廠 (ISO)"]

vc_eval_hdr = ["頻段", "Leq (μm/s)"] + VC_ORDER_L[:5]
for c, h in enumerate(vc_eval_hdr, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, len(vc_eval_hdr))
r4 += 1

vc_freqs_num = [4,5,6.3,8,10,12.5,16,20,25,31.5,40,50,63,80]
for fn, fv in zip(vc_bands, vc_freqs_num):
    st = RD["vel_stats"].get(fn, {})
    leq_v = st.get("Leq", 0)
    ws4.cell(row=r4, column=1, value=fn).font = cat_font
    ws4.cell(row=r4, column=1).alignment = center
    ws4.cell(row=r4, column=1).border = border
    cell = ws4.cell(row=r4, column=2, value=f"{leq_v:.4f}")
    cell.font = bold_val
    cell.alignment = center
    cell.border = border
    for k, grade_name in enumerate(VC_ORDER_L[:5]):
        flat = VC_CURVES_L[grade_name]
        if fv >= 8:
            limit = flat
        else:
            steps = round(np.log2(8.0 / fv) * 3)
            limit = flat * 10 ** (steps * 2 / 20)
        passed = leq_v <= limit
        cell = ws4.cell(row=r4, column=3 + k,
                        value=f"{limit:.2f}" + (" ✓" if passed else " ✗"))
        cell.font = green_font if passed else Font(name="Consolas", size=10, color="C53030")
        cell.alignment = center
        cell.border = border
    r4 += 1
r4 += 1

# VC grade summary
r4 = subtitle_cell(ws4, r4, "VC 等級評定結果（各統計量）", 10)
grade_sum_hdr = ["統計量", "峰值頻段速度 (μm/s)", "VC 等級"]
for c, h in enumerate(grade_sum_hdr, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, 3)
r4 += 1

for sk in STATS:
    band_vels = []
    for fn, fv in zip(vc_bands, vc_freqs_num):
        v = RD["vel_stats"].get(fn, {}).get(sk, 0)
        if v and v > 0:
            band_vels.append((fv, v))
    grade = "超過工廠 (ISO)"
    for gn in VC_ORDER_L:
        flat = VC_CURVES_L[gn]
        if all(vel <= (flat if fv >= 8 else flat * 10 ** (round(np.log2(8.0 / fv) * 3) * 2 / 20))
               for fv, vel in band_vels):
            grade = gn
            break
    peak = max(v for _, v in band_vels) if band_vels else 0
    ws4.cell(row=r4, column=1, value=sk).font = cat_font
    ws4.cell(row=r4, column=1).alignment = center
    ws4.cell(row=r4, column=1).border = border
    cell = ws4.cell(row=r4, column=2, value=f"{peak:.4f}")
    cell.font = val_font
    cell.alignment = center
    cell.border = border
    cell = ws4.cell(row=r4, column=3, value=grade)
    cell.font = green_font
    cell.alignment = center
    cell.border = border
    r4 += 1
r4 += 1

# ── 寬頻統計 ──
r4 = subtitle_cell(ws4, r4, "九、寬頻合成統計（BB = √Σband²，單位 m/s²）", 10)
bb_hdr = ["統計量", "值 (m/s²)"]
for c, h in enumerate(bb_hdr, 1):
    ws4.cell(row=r4, column=c, value=h)
style_hdr(ws4, r4, 2)
r4 += 1

for s in STATS:
    v = RD["broadband"].get(s, 0)
    ws4.cell(row=r4, column=1, value=s).font = cat_font
    ws4.cell(row=r4, column=1).alignment = center
    ws4.cell(row=r4, column=1).border = border
    cell = ws4.cell(row=r4, column=2, value=f"{v:.4e}")
    cell.font = val_font
    cell.alignment = center
    cell.border = border
    r4 += 1

ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 24
ws4.column_dimensions["C"].width = 56
for c in range(4, 22):
    ws4.column_dimensions[get_column_letter(c)].width = 14

# ═══════════════════════════════════════
# Sheet 5: 原始數據說明
# ═══════════════════════════════════════
with open(r"C:\Users\AC717\背景振動量測數據\raw_xyz_samples.json", "r", encoding="utf-8") as _f:
    RAW = json.load(_f)

ws5 = wb.create_sheet("原始數據說明")
r5 = 1

r5 = title_cell(ws5, r5, "原始數據說明 — .RND 檔案結構與 XYZ 合成過程", 10)
r5 += 1

# ── 一、RND 檔案格式說明 ──
r5 = subtitle_cell(ws5, r5, "一、.RND 檔案格式說明", 10)
fmt_rows = [
    ("檔案格式", "CSV（逗號分隔），第 1 列為 \"CSV\" 標記，第 2 列為欄位標頭"),
    ("量測儀器", "加速規（三軸），輸出 1/3 八度音帶頻譜"),
    ("取樣間距", "0.1 秒 / 筆（每秒 10 筆）"),
    ("量測單位", "m/s²（加速度）"),
    ("頻段範圍", "1 Hz ~ 315 Hz（共 26 個 1/3 八度音帶）+ AP(全頻) + APW(加權)"),
    ("軸向", "X、Y、Z 三軸分別記錄，每軸各有完整頻段欄位"),
    ("欄位命名", "X_10 Hz、Y_10 Hz、Z_10 Hz … 依此類推"),
    ("特殊值", "-- = 超出範圍、UN = Under（低於下限）、OL = Over（超出上限）"),
]
for label, val in fmt_rows:
    ws5.cell(row=r5, column=1, value=label).font = cat_font
    ws5.cell(row=r5, column=1).border = border
    ws5.cell(row=r5, column=2, value=val).font = body_font
    ws5.cell(row=r5, column=2).border = border
    ws5.merge_cells(start_row=r5, start_column=2, end_row=r5, end_column=8)
    r5 += 1
r5 += 1

# ── 二、欄位結構示意 ──
r5 = subtitle_cell(ws5, r5, "二、欄位結構示意（每列 = 0.1 秒一筆資料）", 10)
col_demo = ["Address", "Start Time",
            "X_AP", "X_1 Hz", "X_10 Hz", "… X_315 Hz",
            "Y_AP", "Y_1 Hz", "Y_10 Hz", "… Y_315 Hz",
            "Z_AP", "Z_1 Hz", "Z_10 Hz", "… Z_315 Hz"]
for c, h in enumerate(col_demo, 1):
    cell = ws5.cell(row=r5, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = cwrap
    cell.border = border
r5 += 1
demo_vals = ["1", "2026/04/16 10:00:00.0",
             "0.00011", "0.00001", "0.00005", "…",
             "0.00011", "0.00001", "0.00004", "…",
             "0.00021", "0.00000", "0.00008", "…"]
for c, v in enumerate(demo_vals, 1):
    cell = ws5.cell(row=r5, column=c, value=v)
    cell.font = val_font
    cell.alignment = center
    cell.border = border
r5 += 2

# ── 三、XYZ 合成計算（10 Hz 頻段） ──
for band_name in ["10 Hz", "31.5 Hz"]:
    raw_rows = RAW[band_name]
    r5 = subtitle_cell(ws5, r5,
        f"三、原始 X / Y / Z 數據與 XYZ 合成 — {band_name} 頻段（前 20 筆）" if band_name == "10 Hz"
        else f"四、原始 X / Y / Z 數據與 XYZ 合成 — {band_name} 頻段（前 20 筆）", 10)

    raw_hdr = ["筆數", "時間", f"X_{band_name}\n(m/s²)", f"Y_{band_name}\n(m/s²)",
               f"Z_{band_name}\n(m/s²)", "X²", "Y²", "Z²", "X²+Y²+Z²",
               f"XYZ = √(X²+Y²+Z²)\n(m/s²)"]
    for c, h in enumerate(raw_hdr, 1):
        ws5.cell(row=r5, column=c, value=h)
    style_hdr(ws5, r5, len(raw_hdr))
    r5 += 1

    data_start_row = r5  # remember for Leq formula later
    for i, row in enumerate(raw_rows):
        x, y, z = row["X"], row["Y"], row["Z"]
        t_short = row["time"].replace("2026-04-16 ", "")
        rr = r5  # current Excel row
        # Col A: 筆數
        cell = ws5.cell(row=rr, column=1, value=i+1)
        cell.font = cat_font; cell.border = border; cell.alignment = center
        # Col B: 時間
        cell = ws5.cell(row=rr, column=2, value=t_short)
        cell.font = val_font; cell.border = border; cell.alignment = center
        # Col C/D/E: X, Y, Z 原始數值
        for ci, v in [(3, x), (4, y), (5, z)]:
            cell = ws5.cell(row=rr, column=ci, value=v)
            cell.font = val_font; cell.border = border; cell.alignment = center
            cell.number_format = "0.00000"
        # Col F: X² = C^2
        cl = get_column_letter
        cell = ws5.cell(row=rr, column=6, value=f"={cl(3)}{rr}^2")
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell.number_format = "0.00E+00"
        # Col G: Y² = D^2
        cell = ws5.cell(row=rr, column=7, value=f"={cl(4)}{rr}^2")
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell.number_format = "0.00E+00"
        # Col H: Z² = E^2
        cell = ws5.cell(row=rr, column=8, value=f"={cl(5)}{rr}^2")
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell.number_format = "0.00E+00"
        # Col I: X²+Y²+Z² = F+G+H
        cell = ws5.cell(row=rr, column=9, value=f"={cl(6)}{rr}+{cl(7)}{rr}+{cl(8)}{rr}")
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell.number_format = "0.00E+00"
        # Col J: XYZ = SQRT(I)
        cell = ws5.cell(row=rr, column=10, value=f"=SQRT({cl(9)}{rr})")
        cell.font = bold_val; cell.border = border; cell.alignment = center
        cell.number_format = "0.00000000"
        r5 += 1

    # Leq / Lmax / L50 公式示範行
    r5 += 1
    end_row = data_start_row + len(raw_rows) - 1
    jcol = cl(10)  # XYZ column
    demo_formulas = [
        ("Leq（RMS）", f"=SQRT(SUMPRODUCT({jcol}{data_start_row}:{jcol}{end_row}^2)/COUNT({jcol}{data_start_row}:{jcol}{end_row}))"),
        ("Lmax", f"=MAX({jcol}{data_start_row}:{jcol}{end_row})"),
        ("L50（中位數）", f"=PERCENTILE({jcol}{data_start_row}:{jcol}{end_row},0.5)"),
    ]
    for label, formula in demo_formulas:
        ws5.cell(row=r5, column=1, value=label).font = cat_font
        ws5.cell(row=r5, column=1).border = border
        ws5.cell(row=r5, column=1).alignment = center
        cell = ws5.cell(row=r5, column=2, value=f"公式：{formula}")
        cell.font = code_font; cell.border = border
        ws5.merge_cells(start_row=r5, start_column=2, end_row=r5, end_column=6)
        cell = ws5.cell(row=r5, column=7, value=formula)
        cell.font = bold_val; cell.border = border; cell.alignment = center
        cell.number_format = "0.00000000"
        ws5.merge_cells(start_row=r5, start_column=7, end_row=r5, end_column=10)
        r5 += 1
    r5 += 1

# ── 五、從原始數據到統計量的流程 ──
r5 = subtitle_cell(ws5, r5, "五、從 0.1 秒原始數據到統計量的完整流程", 10)

flow_rows = [
    ("步驟 1", "讀取 .RND 檔案", "逐檔讀取 CSV，篩選指定時段（如 10:00~11:00）"),
    ("步驟 2", "清理特殊值", "將 --、UN、OL 視為無效值排除"),
    ("步驟 3", "XYZ 三軸合成", "逐筆、逐頻段計算 XYZ = √(X² + Y² + Z²)"),
    ("步驟 4", "計算統計量", "對 36,000 筆 XYZ 合成值計算 Leq、L5、L10、L50、L90、L95、Lmax"),
    ("步驟 5", "單位轉換", "加速度 → 速度：v = a/(2πf)；加速度 dB：20·log₁₀(a/10⁻⁶)；速度 dB：20·log₁₀(v/2.54×10⁻⁸)"),
    ("步驟 6", "VC 等級評定", "4~80 Hz 各頻段速度(μm/s) 與 VC 曲線限值比對，判定等級"),
]
flow_hdr = ["階段", "名稱", "說明"]
for c, h in enumerate(flow_hdr, 1):
    ws5.cell(row=r5, column=c, value=h)
style_hdr(ws5, r5, 3)
r5 += 1

for step, name, desc in flow_rows:
    ws5.cell(row=r5, column=1, value=step).font = cat_font
    ws5.cell(row=r5, column=1).border = border
    ws5.cell(row=r5, column=1).alignment = center
    ws5.cell(row=r5, column=2, value=name).font = body_font
    ws5.cell(row=r5, column=2).border = border
    ws5.cell(row=r5, column=3, value=desc).font = val_font
    ws5.cell(row=r5, column=3).border = border
    ws5.merge_cells(start_row=r5, start_column=3, end_row=r5, end_column=8)
    r5 += 1
r5 += 1

# ── 六、數據量計算 ──
r5 = subtitle_cell(ws5, r5, "六、數據量對照", 10)
qty_rows = [
    ("量測時間", "1 小時 = 3,600 秒"),
    ("取樣率", "10 筆/秒（0.1 秒間距）"),
    ("總筆數", "3,600 × 10 = 36,000 筆"),
    ("頻段數", "21 個 1/3 八度音帶（1~100 Hz）"),
    ("軸向數", "3 軸（X、Y、Z）"),
    ("原始數據欄位", "21 頻段 × 3 軸 = 63 欄（+ AP、APW 等輔助欄）"),
    ("XYZ 合成後", "21 頻段 × 36,000 筆 = 756,000 個合成值"),
]
for label, val in qty_rows:
    ws5.cell(row=r5, column=1, value=label).font = cat_font
    ws5.cell(row=r5, column=1).border = border
    ws5.cell(row=r5, column=2, value=val).font = val_font
    ws5.cell(row=r5, column=2).border = border
    ws5.merge_cells(start_row=r5, start_column=2, end_row=r5, end_column=5)
    r5 += 1

ws5.column_dimensions["A"].width = 10
ws5.column_dimensions["B"].width = 22
for c in range(3, 11):
    ws5.column_dimensions[get_column_letter(c)].width = 16

# ═══════════════════════════════════════
# Sheet 6: 1秒聚合說明
# ═══════════════════════════════════════
with open(r"C:\Users\AC717\背景振動量測數據\raw_1s_demo.json", "r", encoding="utf-8") as _f:
    RAW1S = json.load(_f)

ws6 = wb.create_sheet("1秒聚合說明")
r6 = 1

r6 = title_cell(ws6, r6, "從 0.1 秒原始數據聚合為 1 秒 RMS — 以 10 Hz 頻段為例", 12)
r6 += 1

# ── 一、說明 ──
r6 = subtitle_cell(ws6, r6, "一、聚合原理說明", 12)
desc_rows = [
    ("原始間距", "0.1 秒 / 筆（每秒 10 筆）"),
    ("聚合間距", "1 秒 / 筆（App 預設「分析間距」）"),
    ("聚合方式", "RMS（均方根）= √(1/N × Σaᵢ²)，N = 10"),
    ("意義", "將每 10 筆 0.1 秒數據壓縮為 1 筆 1 秒等效值"),
    ("公式", "a_1s = √((a₁² + a₂² + … + a₁₀²) / 10)"),
    ("對應關係", "「實測範例」工作表中的 20 筆 XYZ 合成加速度 = 本表每秒 RMS 結果"),
]
for label, val in desc_rows:
    ws6.cell(row=r6, column=1, value=label).font = cat_font
    ws6.cell(row=r6, column=1).border = border
    ws6.cell(row=r6, column=2, value=val).font = body_font
    ws6.cell(row=r6, column=2).border = border
    ws6.merge_cells(start_row=r6, start_column=2, end_row=r6, end_column=8)
    r6 += 1
r6 += 1

cl = get_column_letter

for band_name in ["10 Hz", "31.5 Hz"]:
    raw_rows = RAW1S[band_name]
    sec_label = "二" if band_name == "10 Hz" else "三"
    r6 = subtitle_cell(ws6, r6,
        f"{sec_label}、{band_name} 頻段 — 0.1 秒原始數據（200 筆 = 20 秒）", 12)

    # ── 0.1s 原始數據表 ──
    raw01_hdr = ["筆數", "時間", f"X_{band_name}\n(m/s²)", f"Y_{band_name}\n(m/s²)",
                 f"Z_{band_name}\n(m/s²)", f"XYZ 合成\n= √(X²+Y²+Z²)", "所屬秒數"]
    for c, h in enumerate(raw01_hdr, 1):
        ws6.cell(row=r6, column=c, value=h)
    style_hdr(ws6, r6, len(raw01_hdr))
    r6 += 1

    raw_data_start = r6
    for i, row in enumerate(raw_rows):
        rr = r6
        sec_no = i // 10 + 1
        t_short = row["time"].replace("2026-04-16 ", "")
        # Col A: 筆數
        cell = ws6.cell(row=rr, column=1, value=i+1)
        cell.font = cat_font; cell.border = border; cell.alignment = center
        # Col B: 時間
        cell = ws6.cell(row=rr, column=2, value=t_short)
        cell.font = val_font; cell.border = border; cell.alignment = center
        # Col C/D/E: X, Y, Z
        for ci, v in [(3, row["X"]), (4, row["Y"]), (5, row["Z"])]:
            cell = ws6.cell(row=rr, column=ci, value=v)
            cell.font = val_font; cell.border = border; cell.alignment = center
            cell.number_format = "0.00000"
        # Col F: XYZ = SQRT(C^2+D^2+E^2) 公式
        cell = ws6.cell(row=rr, column=6,
                        value=f"=SQRT({cl(3)}{rr}^2+{cl(4)}{rr}^2+{cl(5)}{rr}^2)")
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell.number_format = "0.00000000"
        # Col G: 所屬秒數
        cell = ws6.cell(row=rr, column=7, value=f"第 {sec_no} 秒")
        cell.font = note_font; cell.border = border; cell.alignment = center
        # 每 10 行交替底色
        if sec_no % 2 == 0:
            alt_fill = PatternFill("solid", fgColor="F0F4F8")
            for cc in range(1, 8):
                ws6.cell(row=rr, column=cc).fill = alt_fill
        r6 += 1
    raw_data_end = r6 - 1
    r6 += 1

    # ── 1秒 RMS 聚合結果 ──
    sec_label2 = "二" if band_name == "10 Hz" else "三"
    r6 = subtitle_cell(ws6, r6,
        f"{sec_label}（續）、{band_name} — 1 秒 RMS 聚合結果（20 筆）", 12)

    rms_hdr = ["秒數", "起始時間", "0.1s 筆數範圍",
               "RMS 公式\n= √(Σaᵢ²/10)", f"1秒 RMS 值\n(m/s²)",
               f"速度轉換\nv = a/(2πf)\n(μm/s)"]
    for c, h in enumerate(rms_hdr, 1):
        ws6.cell(row=r6, column=c, value=h)
    style_hdr(ws6, r6, len(rms_hdr))
    r6 += 1

    f_val = float(band_name.replace(" Hz", ""))
    rms_data_start = r6
    for s in range(20):
        rr = r6
        row0 = raw_data_start + s * 10
        row9 = raw_data_start + s * 10 + 9
        t_sec = raw_rows[s * 10]["time"].replace("2026-04-16 ", "").split(".")[0]
        xyz_col = cl(6)  # F column = XYZ in raw table

        cell = ws6.cell(row=rr, column=1, value=s+1)
        cell.font = cat_font; cell.border = border; cell.alignment = center
        cell = ws6.cell(row=rr, column=2, value=t_sec)
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell = ws6.cell(row=rr, column=3, value=f"第 {s*10+1}~{s*10+10} 筆")
        cell.font = note_font; cell.border = border; cell.alignment = center
        # 公式文字說明
        cell = ws6.cell(row=rr, column=4,
                        value=f"SQRT(SUMSQ({xyz_col}{row0}:{xyz_col}{row9})/10)")
        cell.font = code_font; cell.border = border; cell.alignment = center
        # 實際 RMS 公式
        cell = ws6.cell(row=rr, column=5,
                        value=f"=SQRT(SUMSQ({xyz_col}{row0}:{xyz_col}{row9})/10)")
        cell.font = bold_val; cell.border = border; cell.alignment = center
        cell.number_format = "0.00000000"
        # 速度轉換公式: v(μm/s) = a / (2*PI()*f) * 1000000
        ecol = cl(5)
        cell = ws6.cell(row=rr, column=6,
                        value=f"={ecol}{rr}/(2*PI()*{f_val})*1000000")
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell.number_format = "0.0000"

        if (s+1) % 2 == 0:
            alt_fill2 = PatternFill("solid", fgColor="F0F4F8")
            for cc in range(1, 7):
                ws6.cell(row=rr, column=cc).fill = alt_fill2
        r6 += 1
    rms_data_end = r6 - 1
    r6 += 1

    # ── 1秒 RMS 的統計量 ──
    r6 = subtitle_cell(ws6, r6,
        f"{sec_label}（統計）、{band_name} — 對 20 筆 1 秒 RMS 計算統計量", 12)

    stat_hdr2 = ["統計量", "公式", "結果 (m/s²)", "結果 (μm/s)"]
    for c, h in enumerate(stat_hdr2, 1):
        ws6.cell(row=r6, column=c, value=h)
    style_hdr(ws6, r6, 4)
    r6 += 1

    ecol5 = cl(5)  # 1s RMS column
    stat_formulas = [
        ("Leq (RMS)", f"SQRT(SUMPRODUCT({ecol5}{rms_data_start}:{ecol5}{rms_data_end}^2)/COUNT({ecol5}{rms_data_start}:{ecol5}{rms_data_end}))"),
        ("L5 (P95)",  f"PERCENTILE({ecol5}{rms_data_start}:{ecol5}{rms_data_end},0.95)"),
        ("L10 (P90)", f"PERCENTILE({ecol5}{rms_data_start}:{ecol5}{rms_data_end},0.90)"),
        ("L50 (P50)", f"PERCENTILE({ecol5}{rms_data_start}:{ecol5}{rms_data_end},0.50)"),
        ("L90 (P10)", f"PERCENTILE({ecol5}{rms_data_start}:{ecol5}{rms_data_end},0.10)"),
        ("L95 (P5)",  f"PERCENTILE({ecol5}{rms_data_start}:{ecol5}{rms_data_end},0.05)"),
        ("Lmax",      f"MAX({ecol5}{rms_data_start}:{ecol5}{rms_data_end})"),
    ]
    for sname, sfml in stat_formulas:
        rr = r6
        cell = ws6.cell(row=rr, column=1, value=sname)
        cell.font = cat_font; cell.border = border; cell.alignment = center
        cell = ws6.cell(row=rr, column=2, value=sfml)
        cell.font = code_font; cell.border = border
        # m/s² 結果
        cell = ws6.cell(row=rr, column=3, value=f"={sfml}")
        cell.font = bold_val; cell.border = border; cell.alignment = center
        cell.number_format = "0.00000000"
        # μm/s 轉換
        c3col = cl(3)
        cell = ws6.cell(row=rr, column=4,
                        value=f"={c3col}{rr}/(2*PI()*{f_val})*1000000")
        cell.font = val_font; cell.border = border; cell.alignment = center
        cell.number_format = "0.0000"
        r6 += 1
    r6 += 2

ws6.column_dimensions["A"].width = 10
ws6.column_dimensions["B"].width = 20
ws6.column_dimensions["C"].width = 16
ws6.column_dimensions["D"].width = 50
ws6.column_dimensions["E"].width = 18
ws6.column_dimensions["F"].width = 18
ws6.column_dimensions["G"].width = 12

out = r"S:\noise\Vibration&Noise\project\15002-(寶山用地)第2期擴建-再生水輸水管線工程設計及監造工作(設計部分)\2.工作區\22527\公式說明.xlsx"
wb.save(out)
print(f"OK: {out}")
